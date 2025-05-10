from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from classes.DBType import DBType, DBTypes
from data_to_db.data_to_sql import load_json, write_json

def expand_excel(excel_path: str):
    """
    Expands the columns visually in Excel such that users does not have to manually adjust the width.

    :param excel_path: The path to the Excel file.
    """
    # Load the workbook and get the active sheet
    wb = load_workbook(excel_path)
    ws = wb.active

    # Auto-adjust column width based on the max length in each column
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_length = 0
        column = get_column_letter(col_idx)
        for cell in column_cells:
            try:
                cell_value = str(cell.value)
                if cell_value:
                    max_length = max(max_length, len(cell_value))
            except:
                pass
        adjusted_width = max_length + 2  # add a little padding
        ws.column_dimensions[column].width = adjusted_width

    # Save the updated workbook
    wb.save(excel_path)

def update_query_metrics(db_type: DBType, query_name: str, time: float, memory: float, output_length: int, query_metrics_file_base_name: str):
    """
    Updates the JSON file containing the query times and memory needed for each query for a database.

    :param db_type: Type of database.
    :param query_name: Name of the query.
    :param time: Time needed to execute the query.
    :param memory: The memory (KB) needed to execute the query.
    :param output_length: Length of the output dataframe.
    :param query_metrics_file_base_name: JSON base name for the query times.
    """
    # Do not save if output length is 0, this cannot be a valid result
    if output_length == 0:
        print(f'WARNING: Output length of query {query_name} is 0. Skipping saving.')
        return

    path = f'{query_metrics_file_base_name}_{db_type.get_type().display_name.lower()}_{db_type.name_suffix}.json'
    current_metrics_data = load_json(path)
    if query_name not in current_metrics_data:
        current_metrics_data[query_name] = {}

    # Add time
    if 'times' not in current_metrics_data[query_name]:
        current_metrics_data[query_name]['times'] = []
    current_metrics_data[query_name]['times'].append(round(time, 4))

    # Add memory
    if 'memories' not in current_metrics_data[query_name]:
        current_metrics_data[query_name]['memories'] = []
    current_metrics_data[query_name]['memories'].append(round(memory, 4))

    # Add output length of dataframe
    if 'output_lengths' not in current_metrics_data[query_name]:
        current_metrics_data[query_name]['output_lengths'] = []
    current_metrics_data[query_name]['output_lengths'].append(output_length)

    write_json(current_metrics_data, path)

def get_total_queries_number(json_queries: dict, db_types: list[DBType]) -> int:
    """
    Gets the total number of queries for all query types.

    :param json_queries: JSON containing the queries for each query type.
    :param db_types: List of database types to include in the count.
    :return: Total number of queries.
    """
    total_queries = 0
    for query_type in json_queries.keys():
        queries = json_queries[query_type]['queries']
        total_queries += len(queries)

    # Multiply by the number of databases, because each query is executed on each database
    return total_queries * len(db_types)